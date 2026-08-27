#!/usr/bin/env python3
"""
Converte a TIPI oficial (planilha .xlsx da Receita Federal) para o CSV aceito por
scripts/build_tabelas_db.py --tipi.

Fonte oficial (Decreto 11.158/2022 e atualizações):
    https://www.gov.br/receitafederal/pt-br/acesso-a-informacao/legislacao/documentos-e-arquivos/tipi.xlsx

Uso:
    python scripts/tipi_xlsx_to_csv.py tipi.xlsx tipi.csv
    python scripts/build_tabelas_db.py --tipi tipi.csv

Regras de conversão:
- Só entram linhas com NCM de 8 dígitos (formato 9999.99.99); posições, subposições
  e capítulos (sem alíquota própria) são ignorados.
- A alíquota "NT" (não tributado) vira vazio/NULL; alíquotas numéricas são
  arredondadas a 2 casas (a planilha carrega ruído de ponto flutuante).
- Linhas "Ex" (exceções da TIPI) do mesmo NCM não geram registro próprio: são
  agregadas no campo ``ex_tipi`` como "Ex 01: descrição (alíquota%)".
- Prefixos hierárquicos da descrição ("- ", "-- ") são removidos. Quando a linha de
  8 dígitos traz só um texto genérico ("Outros", "Outras", "Outro(a)s..."), a
  descrição é composta com os ancestrais (posição/subposição/item), cada um
  truncado em 90 caracteres, ex.: "Máquinas automáticas para processamento de
  dados... > Portáteis, de peso não superior a 10 kg > De peso inferior a 3,5 kg > Outras".
- A TIPI não informa unidade tributável; o campo ``unidade_tributavel`` fica vazio.

Requer ``openpyxl`` (``pip install openpyxl``); não é dependência do pacote.
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from collections import OrderedDict
from collections.abc import Iterable, Sequence
from pathlib import Path

_RE_NCM8 = re.compile(r"^\d{4}\.\d{2}\.\d{2}$")
_RE_CODIGO = re.compile(r"^\d{2}(?:\.\d{1,2})*(?:\.\d{1,2})?$|^\d{4}(?:\.\d{1,2})*$")
_COLUNAS = ("codigo", "descricao", "aliquota_ipi", "unidade_tributavel", "ex_tipi")


def _limpar_descricao(texto: object) -> str:
    descricao = " ".join(str(texto or "").split())
    return descricao.lstrip("- ").strip().rstrip(":").strip()


_RE_GENERICO = re.compile(r"^outr[ao]s?\b", re.IGNORECASE)
_MAX_ANCESTRAL = 90


def _truncar(texto: str, limite: int = _MAX_ANCESTRAL) -> str:
    return texto if len(texto) <= limite else texto[: limite - 1].rstrip() + "…"


def _compor_descricao(digitos: str, contexto: dict[str, str], folha: str) -> str:
    """Prefixa a folha com os ancestrais (4 a 7 dígitos) quando ela é genérica."""
    if not _RE_GENERICO.match(folha):
        return folha
    partes: list[str] = []
    for n in (4, 5, 6, 7):
        ancestral = contexto.get(digitos[:n])
        if ancestral and (not partes or partes[-1] != ancestral):
            partes.append(_truncar(ancestral))
    partes.append(folha)
    return " > ".join(partes)


def _normalizar_aliquota(valor: object) -> str:
    texto = str(valor if valor is not None else "").strip().replace(",", ".")
    if not texto or texto.upper() == "NT":
        return ""
    try:
        return f"{round(float(texto), 2):g}"
    except ValueError:
        return ""


def converter_linhas(linhas: Iterable[Sequence[object]]) -> tuple[list[dict[str, str]], int]:
    """Converte as linhas da planilha (tuplas NCM, EX, DESCRIÇÃO, ALÍQUOTA) em registros.

    Retorna (registros no formato do CSV, total de exceções "Ex" agregadas).
    Não depende de openpyxl: recebe qualquer iterável de tuplas/listas.
    """
    registros: OrderedDict[str, dict[str, str]] = OrderedDict()
    excecoes: dict[str, list[str]] = {}
    contexto: dict[str, str] = {}  # dígitos do código (4-7) -> descrição do ancestral
    for linha in linhas:
        if not linha or linha[0] is None:
            continue
        ncm_txt = str(linha[0]).strip()
        ex = linha[1] if len(linha) > 1 else None
        descricao = _limpar_descricao(linha[2] if len(linha) > 2 else "")
        if not _RE_NCM8.match(ncm_txt):
            digitos = ncm_txt.replace(".", "")
            if ex in (None, "") and digitos.isdigit() and 4 <= len(digitos) <= 7:
                contexto[digitos] = descricao
                # Um novo ancestral invalida os descendentes anteriores desse prefixo.
                for chave in [
                    k for k in contexto if len(k) > len(digitos) and k.startswith(digitos)
                ]:
                    del contexto[chave]
            continue
        codigo = ncm_txt.replace(".", "")
        aliquota = _normalizar_aliquota(linha[3] if len(linha) > 3 else "")

        if ex not in (None, ""):
            ex_num = str(ex).strip()
            ex_num = ex_num.zfill(2) if ex_num.isdigit() else ex_num
            sufixo = f" ({aliquota}%)" if aliquota else " (NT)"
            excecoes.setdefault(codigo, []).append(f"Ex {ex_num}: {descricao}{sufixo}")
            continue

        if codigo in registros:  # primeira ocorrência prevalece
            continue
        registros[codigo] = {
            "codigo": codigo,
            "descricao": _compor_descricao(codigo, contexto, descricao),
            "aliquota_ipi": aliquota,
            "unidade_tributavel": "",
            "ex_tipi": "",
        }

    for codigo, lista in excecoes.items():
        if codigo in registros:
            registros[codigo]["ex_tipi"] = " | ".join(lista)
    return list(registros.values()), sum(len(v) for v in excecoes.values())


def escrever_csv(registros: Iterable[dict[str, str]], csv_saida: Path) -> None:
    csv_saida.parent.mkdir(parents=True, exist_ok=True)
    with csv_saida.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=_COLUNAS)
        writer.writeheader()
        writer.writerows(registros)


def converter(xlsx: Path, csv_saida: Path) -> tuple[int, int]:
    try:
        import openpyxl
    except ImportError:  # pragma: no cover - dependência opcional
        sys.exit("openpyxl não instalado: pip install openpyxl")

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    registros, total_ex = converter_linhas(ws.iter_rows(values_only=True))
    escrever_csv(registros, csv_saida)
    return len(registros), total_ex


def main() -> None:
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument("xlsx", type=Path, help="Planilha tipi.xlsx da Receita Federal")
    parser.add_argument("csv", type=Path, help="CSV de saída para build_tabelas_db.py --tipi")
    args = parser.parse_args()
    if not args.xlsx.exists():
        sys.exit(f"Erro: arquivo não encontrado: {args.xlsx}")
    total, total_ex = converter(args.xlsx, args.csv)
    print(f"NCM exportados: {total} (exceções Ex agregadas: {total_ex}) -> {args.csv}")


if __name__ == "__main__":
    main()
